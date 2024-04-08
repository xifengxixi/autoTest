import os
import re
import smtplib
import string
from configparser import ConfigParser
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import sys
import time
import configobj
import pytest
function_path = os.path.abspath('.')
function_path = function_path
function_path1 = function_path.replace("api_frame","")
sys.path.append(function_path)
sys.path.append(function_path1)
from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import Workbook
from api_frame.config import ApiConfig

def failed_excel(read_report, save_path):
    """
    失败的用例信息写入excel
    read_report：读取的html报告
    save_path：存储excel的位置
    """
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, value='用例结果')
    ws.cell(1, 2, value='用例描述')
    ws.cell(1, 3, value='用例位置')
    ws.cell(1, 4, value='运行时间')
    ws.cell(1, 5, value='失败原因')
    soup = BeautifulSoup(open(read_report), 'html.parser')
    fail_list = soup.find_all(class_='failed results-table-row')
    index = 2
    for i in fail_list:
        ws.cell(index, 1, value=i.find(class_='col-result').string)
        ws.cell(index, 2, value=i.find('td').find_next_sibling('td').string)
        ws.cell(index, 3, value=i.find(class_='col-name').string)
        ws.cell(index, 4, value=i.find(class_='col-duration').string)
        ws.cell(index, 5, value=i.find(class_='log').text)
        index = index + 1
    wb.save(save_path)


def send_mail(file_new, receiver = 'xxxxxxx@qq.com'):
    '''
    发送测试报告到邮箱
    :param file_new: 需要发送的文件路径
    :param receiver: 邮件接收人
    :return:
    '''
    # ----------------------------------------------------------
    # 获取邮件正文,读取测试报告的内容
    f = open(file_new, 'rb')
    mail_body = f.read()
    f.close()
    # 邮件服务器
    smtpserver = 'smtp.163.com'
    # 发件人和密码
    sender = 'xxxxxxx@163.com'
    password = 'xxxxxxx'
    # 接收人
    receiver = receiver
    # 邮件主题
    subject = u'自动化报告'
    # ----------------------------------------------------------
    # 连接登录邮箱
    server = smtplib.SMTP(smtpserver, 25)
    server.login(sender, password)
    # ----------------------------------------------------------
    # 添加附件
    send_file = open(file_new, 'rb').read()
    att = MIMEText(send_file, "base64", 'utf-8')
    att['Content-Type'] = 'application/octet-stream'
    att['Content-Disposition'] = 'attachment;filename="report.html"'
    msg = MIMEMultipart('related')
    msgText = MIMEText(mail_body, 'html', 'utf-8')
    msg.attach(msgText)
    msg['From'] = sender
    msg['To'] = receiver
    msg['Subject'] = Header(subject, 'utf-8').encode()
    msg.attach(att)
    # ----------------------------------------------------------
    # 发送邮件
    # server.sendmail(sender, [receiver], msg.as_string())
    # server.quit()
    # print("email has send out!")

if __name__ =="__main__":
    now = time.strftime("%Y%m%d%H%M%S")
    path = os.getcwd()
    pathCase = path + "/testCase"
    ApiConfig.filename = "account.txt"
    url = "www.testingedu.com.cn:8000"
    module = "task"
    test_list = [
        'test_task_AAA_case',
        'test_task_AAB_case',
    ]
    test_str = ' or '.join(test_list)

    if "test" in str(module) and "dev" in str(url):
        ApiConfig.timeout = 60
    else:
        ApiConfig.timeout = 15
    pathReport = path + "/reports/report" + str(now) + ".html"

    pytest.main([
        pathCase,
        # "-m task and not openapi",
        # r'C:\xfxx\git\autoTest\api_frame\testCase',
        # '-k', test_str,
        # "-m" + str(module),
        "--html=" + pathReport,
        "--self-contained-html",
        #"--full-trace",
        "--reruns", "1"
    ])

